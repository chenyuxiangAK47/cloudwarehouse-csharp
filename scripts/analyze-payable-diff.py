"""Compare bill detail X column payable vs cost table; hunt for 0.2 origin."""
from openpyxl import load_workbook
import os
import sys

def bracket_for_weight(w):
    if w <= 0.3:
        return (0, 0.3)
    if w <= 0.5:
        return (0.3, 0.5)
    if w <= 1:
        return (0.5, 1)
    if w <= 2:
        return (1, 2)
    if w <= 3:
        return (2, 3)
    if w <= 4:
        return (3, 4)
    if w <= 5:
        return (4, 5)
    return None


def read_cost_map(wb, sheet_name):
    ws = wb[sheet_name]
    cost = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        express, prov = str(row[0]), row[2]
        if not prov:
            continue
        try:
            mn, mx = float(row[3]), float(row[4])
        except (TypeError, ValueError):
            continue
        fee = None
        for idx in (7, 10, 13):
            if idx < len(row) and isinstance(row[idx], (int, float)):
                fee = float(row[idx])
                break
        if fee is None:
            continue
        cost[(express, str(prov), mn, mx)] = fee
    return cost


def scan_sheet_for_02(ws, max_row=300):
    hits = []
    for row in ws.iter_rows(max_row=max_row):
        for c in row:
            v = c.value
            if v is None:
                continue
            s = str(v)
            if "0.2" in s or "0.20" in s:
                hits.append((c.coordinate, s[:100]))
    return hits


def analyze_book(path, label):
    print("\n" + "=" * 70)
    print(label, os.path.basename(path))

    wb = load_workbook(path, data_only=False)
    wbv = load_workbook(path, data_only=True)

    cost_name = next((n for n in wb.sheetnames if "成本" in n), None)
    bill_name = next((n for n in wb.sheetnames if n.startswith("2026-01")), None)
    param_sheets = [
        n
        for n in wb.sheetnames
        if "参数" in n or n in ("客户编号", "账户-客户", "账单统计-2026", "导出摘要")
    ]
    print("cost:", cost_name, "| bill:", bill_name)
    print("param-like sheets:", param_sheets)

    for ps in param_sheets:
        hits = scan_sheet_for_02(wb[ps])
        if hits:
            print(f"  [{ps}] cells mentioning 0.2: {len(hits)}")
            for h in hits[:8]:
                print("   ", h)

    if not cost_name or not bill_name:
        wb.close()
        wbv.close()
        return

    cost = read_cost_map(wbv, cost_name)
    expresses = sorted(set(k[0] for k in cost))
    print("cost express count:", len(expresses))
    print("sample expresses:", expresses[:8])

    ws = wb[bill_name]
    wsv = wbv[bill_name]

    # Formula on payable transit fee (col X = 24)
    for r in (3, 4, 5):
        print(f"  X{r} formula: {ws.cell(r, 24).value}")

    stats = {}
    matched_alt = {}
    no_cost = 0
    samples = []

    for r in range(3, wsv.max_row + 1):
        express = wsv.cell(r, 3).value
        prov = wsv.cell(r, 7).value
        rnd = wsv.cell(r, 11).value
        pay = wsv.cell(r, 24).value
        recv = wsv.cell(r, 12).value
        if not prov or pay is None or not isinstance(pay, (int, float)):
            continue
        try:
            rw = float(rnd)
        except (TypeError, ValueError):
            continue
        bk = bracket_for_weight(rw)
        if not bk:
            continue
        mn, mx = bk
        cs = cost.get((express, str(prov), mn, mx))
        if cs is None:
            no_cost += 1
            continue
        d = round(float(pay) - cs, 2)
        stats[d] = stats.get(d, 0) + 1

        if len(samples) < 8 and abs(d - 0.2) < 0.001:
            alts = {
                e: cost.get((e, str(prov), mn, mx))
                for e in expresses
                if cost.get((e, str(prov), mn, mx)) is not None
            }
            samples.append(
                dict(row=r, prov=prov, rw=rw, pay=pay, recv=recv, express=express, cost=cs, diff=d, alts=alts)
            )

        if abs(d) >= 0.001:
            for e in expresses:
                v = cost.get((e, str(prov), mn, mx))
                if v is not None and abs(v - pay) < 0.001:
                    matched_alt[e] = matched_alt.get(e, 0) + 1
                    break

    print("\nPay(X) - Cost(same express, same bracket) distribution:")
    for k, v in sorted(stats.items(), key=lambda x: -x[1])[:12]:
        print(f"  diff {k:+.2f}: {v} rows")
    print("no cost row for express+prov+bracket:", no_cost)
    if matched_alt:
        print("when diff!=0, pay equals OTHER express cost:")
        for e, c in sorted(matched_alt.items(), key=lambda x: -x[1]):
            print(f"  {e}: {c} rows")

    print("\nSample +0.2 rows:")
    for s in samples:
        print(
            f"  r{s['row']} {s['prov']} rw={s['rw']} pay={s['pay']} "
            f"cost({s['express']})={s['cost']} recv={s['recv']}"
        )
        exact = [(e, v) for e, v in s["alts"].items() if abs(v - s["pay"]) < 0.001]
        near = [(e, v, s["pay"] - v) for e, v in s["alts"].items() if abs(s["pay"] - v - 0.2) < 0.001]
        if exact:
            print("    pay EXACTLY matches cost of:", exact)
        if near:
            print("    pay = cost + 0.2 for:", near[:5])

    # Compare 圆通 vs 中通 for Yunnan 0-0.3
    for prov in ("云南省", "河北省", "上海"):
        row = [
            (e, cost.get((e, prov, 0, 0.3)))
            for e in expresses
            if cost.get((e, prov, 0, 0.3)) is not None
        ]
        if row:
            print(f"\n{prov} bracket 0-0.3 costs:", row)

    wb.close()
    wbv.close()


root = r"d:\tools\cloudwarehouse-csharp\excel"
analyze_book(os.path.join(root, "[93] 小二小店-账单统计-2026年.xlsx"), "93")
analyze_book(os.path.join(root, "[新客户]账单统计-2026年.xlsx"), "新客户")
